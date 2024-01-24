#!/usr/bin/env python3

import argparse
import csv
import json
import os
import re
import sys
import shutil
from datetime import datetime
from openpyxl import load_workbook
from numpy import *

# offset by 3
met_index = {
    "PSNR": 0,
    "PSNRHVS": 1,
    "SSIM": 2,
    "FASTSSIM": 3,
    "CIEDE2000": 4,
    "PSNR Cb": 5,
    "PSNR Cr": 6,
    "APSNR": 7,
    "APSNR Cb": 8,
    "APSNR Cr": 9,
    "MSSSIM": 10,
    "Encoding Time": 11,
    "VMAF_old": 12,
    "Decoding Time": 13,
    "PSNR Y (libvmaf)": 14,
    "PSNR Cb (libvmaf)": 15,
    "PSNR Cr (libvmaf)": 16,
    "CIEDE2000 (libvmaf)": 17,
    "SSIM (libvmaf)": 18,
    "MS-SSIM (libvmaf)": 19,
    "PSNR-HVS Y (libvmaf)": 20,
    "PSNR-HVS Cb (libvmaf)": 21,
    "PSNR-HVS Cr (libvmaf)": 22,
    "PSNR-HVS (libvmaf)": 23,
    "VMAF": 24,
    "VMAF-NEG": 25,
    "APSNR Y (libvmaf)": 26,
    "APSNR Cb (libvmaf)": 27,
    "APSNR Cr (libvmaf)": 28,
    "CAMBI (libvmaf)": 29,
    "Enc MD5": 30,
    "Enc Instr Cnt": 31,
    "Enc Cycle Cnt": 32,
    "Dec Instr Cnt": 33,
    "Dec Cycle Cnt": 34,
}

# row_id for different sets inside template.
start_rows = {
    'A1': 2,
    'A2': 50,
    'A3': 182,
    'A4': 230,
    'A5': 266,
    'B1': 290,
    'B2': 350,
    'G1': 416,
    'G2': 440,
    'E': 482,
    'F1': 2,
    'F2': 170
}

# CTC Configs
# LD : ctc_sets_mandatory - A1-4K
# RA: ctc_sets_mandatory  + ctc_sets_optional
# AI: ctc_sets_mandatory_ai + ctc_sets_optional
# AS: A1 with Downsampling
ctc_sets_mandatory = [
    "aomctc-a1-4k",
    "aomctc-a2-2k",
    "aomctc-a3-720p",
    "aomctc-a4-360p",
    "aomctc-a5-270p",
    "aomctc-b1-syn",
    "aomctc-b2-syn"]
ctc_sets_mandatory_ai = ctc_sets_mandatory + \
    ["aomctc-f1-hires", "aomctc-f2-midres"]
ctc_sets_mandatory_ld = [x for x in ctc_sets_mandatory if x != 'aomctc-a1-4k']
ctc_sets_optional = ["aomctc-g1-hdr-4k",
                     "aomctc-g2-hdr-2k", "aomctc-e-nonpristine"]

run_cfgs = ['RA', 'LD', 'AI', 'AS']

quality_presets = {
    "daala": [7, 11, 16, 25, 37],
    "x264": list(range(1, 52, 5)),
    "x265": list(range(5, 52, 5)),
    "x265-rt": list(range(5, 52, 5)),
    "xvc": [20, 25, 30, 35, 40],
    "vp8": list(range(12, 64, 5)),
    "vp9": [20, 32, 43, 55, 63],
    "vp9-rt": [20, 32, 43, 55, 63],
    "vp10": [8, 20, 32, 43, 55, 63],
    "vp10-rt": [8, 20, 32, 43, 55, 63],
    "av1": [20, 32, 43, 55, 63],
    "av1-rt": [20, 32, 43, 55, 63],
    "av2-ai": [85, 110, 135, 160, 185, 210],
    "av2-ra": [110, 135, 160, 185, 210, 235],
    "av2-ra-st": [110, 135, 160, 185, 210, 235],
    "av2-ld": [110, 135, 160, 185, 210, 235],
    "av2-as": [110, 135, 160, 185, 210, 235],
    "av2-as-st": [110, 135, 160, 185, 210, 235],
    "av2-f": [60, 85, 110, 135, 160, 185],
    "thor": list(range(7, 43, 3)),
    "thor-rt": list(range(7, 43, 3)),
    "rav1e": [20 * 4, 32 * 4, 43 * 4, 55 * 4, 63 * 4],
    "svt-av1": [27, 33, 39, 46, 52, 58],
    "svt-av1-ra": [27, 33, 39, 46, 52, 58],
    "svt-av1-ra-crf": [27, 33, 39, 46, 52, 58],
    "svt-av1-ra-vbr": [250, 500, 1000, 4000, 8000, 12000],
    "svt-av1-ra-vbr-2p": [250, 500, 1000, 4000, 8000, 12000],
    "svt-av1-ld-cbr": [250, 500, 1000, 4000, 8000, 12000],
    "svt-av1-ra-cq": [27, 33, 39, 46, 52, 58],
    "svt-av1-as": [27, 33, 39, 46, 52, 58],
    "svt-av1-as-ctc": [23, 27, 31, 35, 39, 43, 47, 51, 55, 59, 63],
    "vvc-vtm": [22, 27, 32, 37],
    "vvc-vtm-ra": [22, 27, 32, 37],
    "vvc-vtm-ra-ctc": [22, 27, 32, 37, 42, 47],
    "vvc-vtm-as-ctc": [22, 27, 32, 37, 42, 47],
    "vvc-vtm-ra-st": [22, 27, 32, 37],
    "vvc-vtm-ld": [22, 27, 32, 37],
    "vvc-vtm-ai": [22, 27, 32, 37],
}

row_header = [
    "TestCfg",
    "EncodeMethod",
    "CodecName",
    "EncodePreset",
    "Class",
    "Name",
    "OrigRes",
    "FPS",
    "BitDepth",
    "CodedRes",
    "QP",
    "Bitrate(kbps)",
    "PSNR_Y",
    "PSNR_U",
    "PSNR_V",
    "SSIM_Y(dB)",
    "MS-SSIM_Y(dB)",
    "VMAF_Y",
    "VMAF_Y-NEG",
    "PSNR-HVS",
    "CIEDE2000",
    "APSNR_Y",
    "APSNR_U",
    "APSNR_V",
    "CAMBI",
    "EncT[s]",
    "DecT[s]",
    "EncInstr",
    "DecInstr",
    "EncCycles",
    "DecCycles",
    "EncMD5"
]

# AS Header is different, so dynamically generating removing the cols
row_header_as = [x for x in row_header if x not in ['CAMBI']]


class Logger(object):
    def __init__(self, run_path, args):
        self.this_args = args
        if not args.ctc_export:
            self.terminal = sys.stdout
        self.log = open(run_path + "/csv_export.csv", "w")

    def write(self, message):
        if not self.this_args.ctc_export:
            self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        if not self.this_args.ctc_export:
            self.terminal.flush()
        self.log.flush()


def update_start_rows(ctc_version, start_rows):
    # Multiple clips are removed/moved for AOM-CTCv5.0
    if ctc_version >= 5.0:
        start_rows['A3'] = 164
        start_rows['A4'] = 212
        start_rows['A5'] = 248
        start_rows['B1'] = 272
        start_rows['B2'] = 332
        start_rows['G1'] = 386
        start_rows['G2'] = 410
        start_rows['E'] = 452


def return_start_rows(set_name, run_info):
    try:
        if 'aomctc' in set_name:
            normalized_set = set_name.split('-')[1].upper()
            if 'ctcVersion' in run_info:
                this_ctc_version = float(run_info['ctcVersion'])
            else:
                this_ctc_version = 4.0
            # Update the Row IDs for CTCv5.0
            update_start_rows(this_ctc_version, start_rows)
            if normalized_set in start_rows.keys():
                return start_rows[normalized_set], normalized_set
    except BaseException as e:
        print(
            "DEBUG: E: return_start_rows: Not a CTC set to Normalize, check the runs,",
            e)
        sys.exit(1)


def return_ctc_set_list(run_info, config):
    run_set_list = []
    # TODO: Extend multipreset to other codecs, but catch this edge case for
    # now
    if len(run_info['ctcSets']) > 0:
        set_name = run_info['ctcSets']
        if 'aomctc-all' in set_name and ('av2' in config or 'vvc' in config):
            if config in ['av2-ai', 'vvc-vtm-ai']:
                run_set_list = ctc_sets_mandatory_ai + ctc_sets_optional
            elif config in ['av2-ra-st', 'av2-ra', 'vvc-vtm-ra', 'vvc-vtm-ra-st', 'vvc-vtm-ra-ctc']:
                run_set_list = ctc_sets_mandatory + ctc_sets_optional
            elif config in ['av2-ld', 'vvc-vtm-ld']:
                run_set_list = ctc_sets_mandatory_ld
            else:
                run_set_list = [run_info['task']]
        elif 'aomctc-mandatory' in set_name and ('av2' in config or 'vvc' in config):
            if config in ['av2-ra-st', 'av2-ra', 'vvc-vtm-ra',
                          'vvc-vtm-ra-st', 'vvc-vtm-ra-ctc']:
                run_set_list = ctc_sets_mandatory
            elif config in ['av2-ld', 'vvc-vtm-ld']:
                run_set_list = ctc_sets_mandatory_ld
            elif config in ['av2-ai', 'vvc-vtm-ai']:
                run_set_list = ctc_sets_mandatory_ai
            else:
                run_set_list = [run_info['task']]
        # Case of AOM Sets but not AVM/VVC codec
        elif 'av2' not in config and 'vvc' not in config and 'rav1e' not in config:
            run_set_list = [run_info['task']]
        else:
            run_set_list = run_info['ctcSets']
    else:
        run_set_list = [run_info['task']]
    return run_set_list


def return_ctc_cfg_list(run_info):
    try:
        cfg_name = run_info['ctcPresets']
        if len(cfg_name) > 0:
            if 'av2-all' in cfg_name:
                return ['av2-ra-st', 'av2-ai', 'av2-ld']
            else:
                return cfg_name
        else:
            return [run_info['codec']]
    except BaseException as e:
        print("DEBUG: W: return_ctc_cfg_list", e)
        return [run_info['codec']]


def return_config(this_config):
    if 'av2-' in this_config:
        if this_config.split('-')[1].upper() in run_cfgs:
            this_cfg = this_config.split('-')[1].upper()
    # VVC has VVC-VTM-* style of naming presets!
    elif 'vvc-' in this_config:
        if this_config.split('-')[2].upper() in run_cfgs:
            this_cfg = this_config.split('-')[2].upper()
    else:
        this_cfg = 'RA'
    return this_cfg


def write_set_data(run_path, writer, current_video_set, current_config):
    info_data = json.load(open(run_path + "/info.json"))
    videos_dir = os.path.join(
        os.getenv("MEDIAS_SRC_DIR", "/mnt/runs/sets"), current_video_set
    )  # for getting framerate
    sets = json.load(
        open(os.path.join(os.getenv("CONFIG_DIR", "rd_tool"), "sets.json")))
    videos = sets[current_video_set]["sources"]
    # sort name ascending, resolution descending
    if current_video_set == "aomctc-a1-4k-as":
        videos.sort(
            key=lambda x: x.split("_")[0]
            + "%08d" % (100000 - int(x.split("_")[1].split("x")[0]))
        )
    # Case of normal CTC clips apart from AS,
    # Deprecrate sorting for CTC reasons now!!
    else:
        pass
    # CTCv5: Obtain CTC version
    if 'ctcVersion' in info_data:
        ctc_version = float(info_data['ctcVersion'])
    else:
        ctc_version = 4.0
    # CTCv4.0: Rollback to CTCv4.0
    if ctc_version < 5.0:
        if current_video_set in ['aomctc-a2-2k',
                                 'aomctc-b2-syn', 'aomctc-e-nonpristine']:
            videos = sets[current_video_set]['CTC_4.0']
    if 'av2' in info_data['codec']:
        normalized_cfg = info_data['codec'].split('-')[1].upper()
    elif 'vvc' in info_data['codec']:
        try:
            normalized_cfg = info_data['codec'].split('-')[2].upper()
        except:
            normalized_cfg = ''
    else:
        normalized_cfg = 'RA'
    # Get the Quality values, if user defined, use that, else do defaults
    if 'qualities' in list(info_data.keys()):
        qp_list = [int(x) for x in info_data['qualities'].split()]
    else:
        qp_list = quality_presets[info_data['codec']]
        if current_config == 'av2-ai':
            if current_video_set in ['aomctc-f1-hires', 'aomctc-f2-midres']:
                qp_list = quality_presets['av2-f']
            else:
                qp_list = quality_presets['av2-ai']
    try:
        for video in videos:
            v = open(os.path.join(videos_dir, video), "rb")
            line = v.readline().decode("utf-8")
            fps_n, fps_d = re.search(r"F([0-9]*)\:([0-9]*)", line).group(1, 2)
            width = re.search(r"W([0-9]*)", line).group(1)
            height = re.search(r"H([0-9]*)", line).group(1)
            if 'aomctc' in current_video_set:
                normalized_set = current_video_set.split('-')[1].upper()
                normalized_cfg = return_config(current_config)
            else:
                normalized_set = ''
            daala_path = os.path.join(
                run_path,
                current_video_set,
                video +
                "-daala.out")
            if 'ctcPresets' in info_data.keys():
                if len(info_data['ctcPresets']
                       ) > 1 or 'av2-all' in info_data['ctcPresets']:
                    daala_path = os.path.join(
                        run_path,
                        current_config,
                        current_video_set,
                        video +
                        "-daala.out")
            a = genfromtxt(daala_path, dtype=None, encoding=None)
            # This way, even partial information from the *daala.out can be
            # rendered by having key-value where key is QP.
            encoded_qp_list = {}
            for row in a:
                encoded_qp_list[int(row[0])] = row
            for this_qp in qp_list:
                # Check if the QPs is present in the currently stored daala.out
                if this_qp in encoded_qp_list.keys():
                    row = encoded_qp_list[this_qp]
                    frames = int(row[1]) / int(width) / int(height)
                    # For still-images set compute Bytes
                    if normalized_set in ['F1', 'F2']:
                        enc_bitrate = round(int(row[2]) * 8.0, 6)
                    else:
                        enc_bitrate = round(
                            int(row[2]) * 8.0 * float(fps_n) / float(fps_d) / frames / 1000.0, 6)
                    enc_md5 = ''
                    enc_instr_cnt = 0
                    enc_cycle_cnt = 0
                    dec_instr_cnt = 0
                    dec_cycle_cnt = 0
                    if len(row) > 33:
                        enc_md5 = str(row[met_index["Enc MD5"] + 3])
                        enc_instr_cnt = int(
                            row[met_index["Enc Instr Cnt"] + 3])
                        enc_cycle_cnt = int(
                            row[met_index["Enc Cycle Cnt"] + 3])
                        dec_instr_cnt = int(
                            row[met_index["Dec Instr Cnt"] + 3])
                        dec_cycle_cnt = int(
                            row[met_index["Dec Cycle Cnt"] + 3])
                    if info_data["codec"] in [
                            "av2-as", "av2-as-st", 'vvc-vtm-as-ctc']:
                        writer.writerow(
                            [
                                "AS",  # TestCfg
                                "aom",  # EncodeMethod
                                info_data["run_id"],  # CodecName
                                "",  # EncodePreset
                                normalized_set,  # Class
                                video,  # name
                                "3840x2160",  # OrigRes
                                str(float(fps_n) / float(fps_d)),  # FPS
                                10,  # BitDepth
                                str(width) + "x" + str(height),  # CodedRes
                                row[0],  # qp
                                enc_bitrate,  # bitrate
                                row[met_index["PSNR Y (libvmaf)"] + 3],
                                row[met_index["PSNR Cb (libvmaf)"] + 3],
                                row[met_index["PSNR Cr (libvmaf)"] + 3],
                                row[met_index["SSIM (libvmaf)"] + 3],
                                row[met_index["MS-SSIM (libvmaf)"] + 3],
                                row[met_index["VMAF"] + 3],
                                row[met_index["VMAF-NEG"] + 3],
                                row[met_index["PSNR-HVS (libvmaf)"] + 3],
                                row[met_index["CIEDE2000 (libvmaf)"] + 3],
                                row[met_index["APSNR Y (libvmaf)"] + 3],
                                row[met_index["APSNR Cb (libvmaf)"] + 3],
                                row[met_index["APSNR Cr (libvmaf)"] + 3],
                                row[met_index["Encoding Time"] + 3],
                                row[met_index["Decoding Time"] + 3],
                                int(enc_instr_cnt),  # EncInstr
                                int(dec_instr_cnt),  # DecInstr
                                int(enc_cycle_cnt),  # EncCycles
                                int(dec_cycle_cnt),  # DecCycles
                                enc_md5,        # EncMD5
                            ]
                        )
                    else:
                        writer.writerow(
                            [
                                normalized_cfg,  # TestCfg
                                "aom",  # EncodeMethod
                                info_data["run_id"],  # CodecName
                                0,  # EncodePreset #TODO: FIXME
                                normalized_set,  # Class
                                video,  # name
                                str(width) + "x" + str(height),  # OrigRes
                                str(float(fps_n) / float(fps_d)),  # FPS
                                10,  # BitDepth #TODO: FIXME
                                str(width) + "x" + str(height),  # CodedRes
                                int(row[0]),  # qp
                                enc_bitrate,  # bitrate
                                row[met_index["PSNR Y (libvmaf)"] + 3],
                                row[met_index["PSNR Cb (libvmaf)"] + 3],
                                row[met_index["PSNR Cr (libvmaf)"] + 3],
                                row[met_index["SSIM (libvmaf)"] + 3],
                                row[met_index["MS-SSIM (libvmaf)"] + 3],
                                row[met_index["VMAF"] + 3],
                                row[met_index["VMAF-NEG"] + 3],
                                row[met_index["PSNR-HVS (libvmaf)"] + 3],
                                row[met_index["CIEDE2000 (libvmaf)"] + 3],
                                row[met_index["APSNR Y (libvmaf)"] + 3],
                                row[met_index["APSNR Cb (libvmaf)"] + 3],
                                row[met_index["APSNR Cr (libvmaf)"] + 3],
                                row[met_index["CAMBI (libvmaf)"] + 3],
                                row[met_index["Encoding Time"] + 3],
                                row[met_index["Decoding Time"] + 3],
                                int(enc_instr_cnt),  # EncInstr
                                int(dec_instr_cnt),  # DecInstr
                                int(enc_cycle_cnt),  # EncCycles
                                int(dec_cycle_cnt),  # DecCycles
                                enc_md5,        # EncMD5
                            ]
                        )
                # Case where the data is yet to be made
                else:
                    writer.writerow([
                        normalized_cfg,  # TestCfg
                        "aom",  # EncodeMethod
                        info_data["run_id"],  # CodecName
                        0,  # EncodePreset #TODO: FIXME
                        normalized_set,  # Class
                        video,  # name
                        str(width) + "x" + str(height),  # OrigRes
                        str(float(fps_n) / float(fps_d)),  # FPS
                        10,  # BitDepth #TODO: FIXME
                        str(width) + "x" + str(height),  # CodedRes
                        this_qp  # qp
                    ])
    except BaseException as e:
        print("DEBUG: W: write_set_data ", e)
        # This allows partial rendering of CSV + XLS Reports
        pass


def save_ctc_export(run_path, cmd_args):
    info_data = json.load(open(run_path + "/info.json"))
    task = info_data["task"]
    sets = json.load(
        open(os.path.join(os.getenv("CONFIG_DIR", "rd_tool"), "sets.json")))
    videos = sets[task]["sources"]
    # sort name ascending, resolution descending
    if task == "aomctc-a1-4k-as":
        videos.sort(
            key=lambda x: x.split("_")[0]
            + "%08d" % (100000 - int(x.split("_")[1].split("x")[0]))
        )
    # Case of normal CTC clips apart from AS,
    # Deprecrate sorting for CTC reasons now!!
    else:
        pass

    ctc_set_list = return_ctc_set_list(info_data, info_data['codec'])
    ctc_cfg_list = return_ctc_cfg_list(info_data)
    cfg_name = info_data['codec']
    if not cmd_args.ctc_export:
        sys.stdout = Logger(run_path, cmd_args)
        csv_writer_obj = sys.stdout
    else:
        csv_writer_obj = open(run_path + "/csv_export.csv", 'w')
    w = csv.writer(csv_writer_obj, dialect="excel")
    if cfg_name in ['av2-as', 'av2-as-st', 'vvc-vtm-as-ctc']:
        w.writerow(row_header_as)
    else:
        w.writerow(row_header)
    # Abstract Writing per-set data
    for config_name in ctc_cfg_list:
        ctc_set_list = return_ctc_set_list(info_data, config_name)
        for set_name in ctc_set_list:
            write_set_data(run_path, w, set_name, config_name)
    # For CTC XLSM case which do not output to STDOUT
    if cmd_args.ctc_export:
        csv_writer_obj.close()


def write_xls_rows(run_path, current_video_set,
                   current_config, this_sheet, this_run_info):
    run_file = open(run_path + '/csv_export.csv', 'r')
    start_id, normalized_set = return_start_rows(
        current_video_set, this_run_info)
    row_end_idx = 31
    if current_config == 'AS':
        row_end_idx = 30
    run_reader = csv.reader(run_file)
    next(run_reader)
    this_row = start_id
    for this_line in run_reader:
        if this_line[0] != current_config or this_line[4] != normalized_set:
            continue
        else:
            this_col = 1
            for this_values in this_line:
                this_cell = this_sheet.cell(row=this_row, column=this_col)
                # QP_val:perf_stats
                if this_col >= 11 and this_col <= row_end_idx and this_values != "":
                    this_cell.value = float(this_values)
                else:
                    this_cell.value = this_values
                this_col += 1
        this_row += 1
    run_file.close()


def write_xls_cfg_sheet(run_a, run_b, run_cfg_list,
                        run_a_info, run_b_info, xls_file, wb):
    # Iterate through CFG lists.
    for cfg_iter in run_cfg_list:
        current_ctc_list_a = return_ctc_set_list(run_a_info, cfg_iter)
        current_ctc_list_b = return_ctc_set_list(run_b_info, cfg_iter)
        this_cfg = return_config(cfg_iter)
        anchor_sheet_name = 'Anchor-%s' % this_cfg
        test_sheet_name = 'Test-%s' % this_cfg
        if this_cfg == 'AS':
            anchor_sheet_name = 'Anchor'
            test_sheet_name = 'Test'
        anchor_sheet = wb[anchor_sheet_name]
        test_sheet = wb[test_sheet_name]
        # Single Video Set Condition
        if len(current_ctc_list_a) == 0 and len(current_ctc_list_b) == 0:
            current_video_set = run_a_info["task"]
            print(
                datetime.isoformat(
                    datetime.now()),
                "Writing data for run",
                run_a,
                "with set",
                current_video_set,
                "and",
                this_cfg,
                "config")
            write_xls_rows(
                run_a,
                current_video_set,
                this_cfg,
                anchor_sheet,
                run_a_info)
            print(
                datetime.isoformat(
                    datetime.now()),
                "Writing data for run",
                run_b,
                "with set",
                current_video_set,
                "and",
                this_cfg,
                "config")
            write_xls_rows(
                run_b,
                current_video_set,
                this_cfg,
                test_sheet,
                run_b_info)
            wb.save(xls_file)
        # Multi-Set Case
        else:
            for this_video_set in current_ctc_list_a:
                print(
                    datetime.isoformat(
                        datetime.now()),
                    "Writing data for run",
                    run_a,
                    "with set",
                    this_video_set,
                    "and",
                    this_cfg,
                    "config")
                if this_video_set in ['aomctc-f1-hires',
                                      'aomctc-f2-midres'] and (cfg_iter == 'av2-ai' or cfg_iter == 'rav1e'):
                    anchor_sheet_name = 'Anchor-Still'
                    anchor_sheet = wb[anchor_sheet_name]
                else:
                    if this_cfg == 'AS':
                        anchor_sheet_name = 'Anchor'
                    else:
                        anchor_sheet_name = 'Anchor-%s' % this_cfg
                    anchor_sheet = wb[anchor_sheet_name]
                write_xls_rows(
                    run_a,
                    this_video_set,
                    this_cfg,
                    anchor_sheet,
                    run_a_info)
            for this_video_set in current_ctc_list_b:
                print(
                    datetime.isoformat(
                        datetime.now()),
                    "Writing data for run",
                    run_b,
                    "with set",
                    this_video_set,
                    "and",
                    this_cfg,
                    "config")
                if this_video_set in ['aomctc-f1-hires',
                                      'aomctc-f2-midres'] and (cfg_iter == 'av2-ai' or cfg_iter == 'rav1e'):
                    test_sheet_name = 'Test-Still'
                    test_sheet = wb[test_sheet_name]
                else:
                    if this_cfg == 'AS':
                        test_sheet_name = 'Test'
                    else:
                        test_sheet_name = 'Test-%s' % this_cfg
                    test_sheet = wb[test_sheet_name]
                write_xls_rows(
                    run_b,
                    this_video_set,
                    this_cfg,
                    test_sheet,
                    run_b_info)
            wb.save(xls_file)


def write_xls_file(run_a, run_b):
    run_a_info = json.load(open(run_a + "/info.json"))
    run_b_info = json.load(open(run_b + "/info.json"))
    run_id_a = run_a_info["run_id"]
    run_id_b = run_b_info["run_id"]
    xls_filename = 'AOM_CWG_Regular_CTCv4_v7.3.2'
    if run_a_info["task"] == 'aomctc-a1-4k-as':
        xls_filename = 'AOM_CWG_AS_CTC_v9.7'
    if ('ctcVersion' in run_a_info) or ('ctcVersion' in run_b_info):
        if float(
                run_a_info['ctcVersion']) >= 5.0:
            xls_filename = 'AOM_CWG_Regular_CTCv5_v7.4.5'
            if run_a_info["task"] == 'aomctc-a1-4k-as':
                xls_filename = 'AOM_CWG_AS_CTC_v9.9'
    xls_template = os.path.join(
        os.getenv("CONFIG_DIR", "rd_tool"), xls_filename + '.xlsm')
    xls_file = run_a + '/../ctc_results/' + \
        "%s-%s-%s.xlsm" % (xls_filename, run_id_a, run_id_b)
    shutil.copyfile(xls_template, xls_file)
    wb = load_workbook(xls_file, read_only=False, keep_vba=True)
    run_a_cfg_list = return_ctc_cfg_list(run_a_info)
    run_b_cfg_list = return_ctc_cfg_list(run_b_info)
    print(datetime.isoformat(datetime.now()), "Start writing data of", run_a)
    write_xls_cfg_sheet(
        run_a,
        run_b,
        run_a_cfg_list,
        run_a_info,
        run_b_info,
        xls_file,
        wb)
    if run_a_cfg_list != run_b_cfg_list:
        print(
            datetime.isoformat(
                datetime.now()),
            "Start writing data of",
            run_b,
            "as config is different.")
        write_xls_cfg_sheet(
            run_a,
            run_b,
            run_b_cfg_list,
            run_a_info,
            run_b_info,
            xls_file,
            wb)


def main():
    parser = argparse.ArgumentParser(
        description="Generate CTC CSV version of .out files")
    parser.add_argument("run", nargs=1, help="Run folder (Anchor)")
    parser.add_argument("--ctc_export", action='store_true', help="XLS Export")
    parser.add_argument(
        "--run_b", help="Target Run Dir (Only used in CTC case)")
    args = parser.parse_args()

    if not args.ctc_export:
        save_ctc_export(args.run[0], args)
    else:
        if not args.run_b:
            print("ERROR: Missing Target, aborting")
            sys.exit(1)
        if args.ctc_export:
            print(
                datetime.isoformat(
                    datetime.now()),
                "CSV Generation for run",
                args.run[0])
        save_ctc_export(args.run[0], args)
        if args.ctc_export:
            print(
                datetime.isoformat(
                    datetime.now()),
                "CSV Generation for run",
                args.run_b)
        save_ctc_export(args.run_b, args)
        if args.ctc_export:
            print(
                datetime.isoformat(
                    datetime.now()),
                "XLSM Generation for run",
                args.run[0],
                "and",
                args.run_b)
        write_xls_file(args.run[0], args.run_b)
        if args.ctc_export:
            print(datetime.isoformat(datetime.now()),
                  "Completed XLSM Generation for the selected jobs")


if __name__ == "__main__":
    main()
